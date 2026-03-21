from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from zipfile import BadZipFile
from datetime import datetime
from collections import Counter
from filter import get_state_label
import os

EXCEL_FILE = "job_tracker.xlsx"

HEADERS = [
    "Title", "Company", "Posted Date", "Days Ago", "State",
    "Score", "Source", "Locations", "Link",
    "Applied", "Priority", "Resume Ver.", "Cover Letter", "Status", "Notes"
]

COL = {h: i + 1 for i, h in enumerate(HEADERS)}

# Source color palette (light tints)
SOURCE_FILLS = {
    "Simplify":   PatternFill("solid", start_color="E8F5E9"),   # mint green
    "Greenhouse": PatternFill("solid", start_color="E3F2FD"),   # sky blue
    "Workday":    PatternFill("solid", start_color="FFF8E1"),   # amber
    "Lever":      PatternFill("solid", start_color="FCE4EC"),   # pink
    "Ashby":      PatternFill("solid", start_color="EDE7F6"),   # purple
    "RemoteOK":   PatternFill("solid", start_color="E0F7FA"),   # cyan
}
ALT_FILL  = PatternFill("solid", start_color="F5F7FA")
EVEN_FILL = PatternFill("solid", start_color="FFFFFF")

# State color overrides (take priority over source colors)
STATE_FILLS = {
    "NJ":     PatternFill("solid", start_color="DCEEFB"),
    "PA":     PatternFill("solid", start_color="D6EFD8"),
    "NY":     PatternFill("solid", start_color="FFF9C4"),
    "Remote": PatternFill("solid", start_color="EDE7F6"),
}

thin = Side(style="thin", color="D8DCE6")
BORDER = Border(bottom=thin)
THICK_BOTTOM = Border(bottom=Side(style="medium", color="B0B8CC"))


def _safe_from_timestamp(ts):
    """Convert unix timestamp to datetime, returning None if out of range."""
    try:
        # Windows only supports timestamps between ~1970 and ~3001
        if ts and 0 < int(ts) < 32503680000:
            return datetime.fromtimestamp(int(ts))
    except Exception:
        pass
    return None


def format_posted_date(job):
    dt = _safe_from_timestamp(job.get("date_updated_ts", 0))
    if dt:
        return dt.strftime("%Y-%m-%d")
    raw = str(job.get("date_updated", "")).strip()
    if raw.isdigit():
        dt = _safe_from_timestamp(raw)
        if dt:
            return dt.strftime("%Y-%m-%d")
    return raw


def days_since_posted(job):
    dt = _safe_from_timestamp(job.get("date_updated_ts", 0))
    if dt:
        return (datetime.now() - dt).days
    raw = str(job.get("date_updated", "")).strip()
    if raw.isdigit():
        dt = _safe_from_timestamp(raw)
        if dt:
            return (datetime.now() - dt).days
    return ""


def load_existing_manual_data(ws):
    manual_data = {}
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col

    required = ["Link", "Applied", "Priority", "Resume Ver.", "Cover Letter", "Status", "Notes"]
    if not all(k in headers for k in required):
        return manual_data

    for row in range(2, ws.max_row + 1):
        url_val = ws.cell(row=row, column=headers["Link"]).value
        if not url_val:
            continue
        url_str = str(url_val).strip()
        if url_str.upper().startswith('=HYPERLINK('):
            try:
                url_str = url_str.split('"')[1]
            except IndexError:
                pass
        manual_data[url_str] = {
            "Applied":       ws.cell(row=row, column=headers["Applied"]).value or "",
            "Priority":      ws.cell(row=row, column=headers["Priority"]).value or "",
            "Resume Ver.":   ws.cell(row=row, column=headers["Resume Ver."]).value or "",
            "Cover Letter":  ws.cell(row=row, column=headers["Cover Letter"]).value or "",
            "Status":        ws.cell(row=row, column=headers["Status"]).value or "",
            "Notes":         ws.cell(row=row, column=headers["Notes"]).value or "",
        }
    return manual_data


def _write_jobs_sheet(wb, jobs, manual_data):
    ws = wb.active
    ws.title = "Jobs"

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", start_color="1A237E")   # deep navy
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_fill and header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = Border(bottom=Side(style="medium", color="FFFFFF"))
    ws.row_dimensions[1].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, job in enumerate(jobs, 2):
        url   = str(job.get("url", "")).strip()
        saved = manual_data.get(url, {})
        state = get_state_label(job)
        source = job.get("source", "")
        score  = job.get("match_score", 0)

        applied_val  = saved.get("Applied",  "") or ("Yes" if job.get("applied") else "No")
        priority_val = saved.get("Priority", job.get("priority", "") or "")
        status_val   = saved.get("Status",   job.get("status",  "") or "")
        notes_val    = saved.get("Notes",    job.get("notes",   "") or "")
        link_formula = f'=HYPERLINK("{url}","↗")' if url else ""

        row_data = [
            job.get("title", ""),
            job.get("company_name", ""),
            format_posted_date(job),
            days_since_posted(job),
            state,
            score,
            source,
            ", ".join(job.get("locations", [])) if isinstance(job.get("locations", []), list) else str(job.get("locations", "")),
            link_formula,
            applied_val, priority_val,
            saved.get("Resume Ver.", ""), saved.get("Cover Letter", ""),
            status_val, notes_val,
        ]

        # Row fill: state takes priority, then source, then alternating
        row_fill = STATE_FILLS.get(state) or SOURCE_FILLS.get(source) or (ALT_FILL if row_idx % 2 == 0 else EVEN_FILL)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border    = BORDER
            cell.fill      = row_fill

            if col_idx == COL["Link"]:
                cell.font      = Font(name="Calibri", size=10, color="1565C0", bold=True)
                cell.alignment = center

            elif col_idx == COL["Score"]:
                cell.alignment = center
                if score >= 12:
                    cell.font = Font(name="Calibri", size=9, bold=True, color="1B5E20")
                    cell.fill = PatternFill("solid", start_color="C8E6C9")
                elif score >= 8:
                    cell.font = Font(name="Calibri", size=9, bold=True, color="E65100")
                    cell.fill = PatternFill("solid", start_color="FFE0B2")
                else:
                    cell.font = Font(name="Calibri", size=9, color="757575")

            elif col_idx == COL["Days Ago"]:
                cell.alignment = center
                try:
                    days = int(value)
                    if days > 21:
                        cell.font = Font(name="Calibri", size=9, color="B71C1C")
                    elif days > 7:
                        cell.font = Font(name="Calibri", size=9, color="E65100")
                except (TypeError, ValueError):
                    pass

            elif col_idx == COL["Applied"]:
                cell.alignment = center
                if str(value).lower() == "yes":
                    cell.font = Font(name="Calibri", size=9, bold=True, color="1B5E20")

            elif col_idx in (COL["State"], COL["Source"], COL["Posted Date"], COL["Priority"]):
                cell.alignment = center

            elif col_idx == COL["Title"] and str(applied_val).lower() == "yes":
                cell.font = Font(name="Calibri", size=9, color="9E9E9E", strike=True)

        ws.row_dimensions[row_idx].height = 16

    # ── Freeze, filter, column widths ─────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    col_widths = {
        "Title": 40, "Company": 22, "Posted Date": 13, "Days Ago": 9,
        "State": 8,  "Score": 7,   "Source": 12,      "Locations": 28,
        "Link": 7,   "Applied": 9, "Priority": 9,     "Resume Ver.": 13,
        "Cover Letter": 13, "Status": 16, "Notes": 32,
    }
    for header, width in col_widths.items():
        ws.column_dimensions[get_column_letter(COL[header])].width = width

    # ── Dropdowns ─────────────────────────────────────────────────────────────
    max_row = ws.max_row

    def add_dv(formula, col_letter):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_row}")

    add_dv('"Yes,No"',                                         get_column_letter(COL["Applied"]))
    add_dv('"1,2,3,4,5"',                                      get_column_letter(COL["Priority"]))
    add_dv('"Applied,Interviewing,Offer,Rejected,Too Lazy"',   get_column_letter(COL["Status"]))

    return ws


def _write_stats_sheet(wb, jobs):
    ws = wb.create_sheet("Stats")
    ws.sheet_view.showGridLines = False

    title_font   = Font(name="Calibri", size=16, bold=True, color="1A237E")
    heading_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    label_font   = Font(name="Calibri", size=10, color="333333")
    value_font   = Font(name="Calibri", size=14, bold=True, color="1A237E")
    hfill        = PatternFill("solid", start_color="1A237E")
    center       = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 16

    ws["A1"] = "Job Tracker — Stats"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=9, color="777777", italic=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 14

    # ── Summary cards ─────────────────────────────────────────────────────────
    now_ts = datetime.now().timestamp()
    week_ago = now_ts - 7 * 86400

    total       = len(jobs)
    applied_n   = sum(1 for j in jobs if str(j.get("applied", "")).lower() in ("yes", "1", "true"))
    new_week    = sum(1 for j in jobs if 0 < (j.get("date_updated_ts") or 0) < 32503680000 and j.get("date_updated_ts") >= week_ago)
    avg_score   = round(sum(j.get("match_score", 0) for j in jobs) / total, 1) if total else 0

    cards = [
        ("Total Jobs",       total,     "A"),
        ("Applied",          applied_n, "C"),
        ("New This Week",    new_week,  "E"),
    ]

    for label, value, col in cards:
        r = 4
        ws[f"{col}{r}"] = label
        ws[f"{col}{r}"].font = heading_font
        ws[f"{col}{r}"].fill = hfill
        ws[f"{col}{r}"].alignment = center
        ws.row_dimensions[r].height = 20

        ws[f"{col}{r+1}"] = value
        ws[f"{col}{r+1}"].font = value_font
        ws[f"{col}{r+1}"].alignment = center
        ws.row_dimensions[r + 1].height = 28

    ws["A6"] = "Avg Score"
    ws["A6"].font = heading_font
    ws["A6"].fill = hfill
    ws["A6"].alignment = center
    ws["A7"] = avg_score
    ws["A7"].font = value_font
    ws["A7"].alignment = center
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 28

    # ── By Source ─────────────────────────────────────────────────────────────
    src_counts = Counter(j.get("source", "Unknown") for j in jobs)
    ws["A9"] = "Jobs by Source"
    ws["A9"].font = Font(name="Calibri", size=11, bold=True, color="1A237E")
    ws.row_dimensions[9].height = 20

    ws["A10"] = "Source"
    ws["B10"] = "Count"
    for cell in [ws["A10"], ws["B10"]]:
        cell.font = heading_font
        cell.fill = hfill
        cell.alignment = center
    ws.row_dimensions[10].height = 18

    for i, (src, cnt) in enumerate(sorted(src_counts.items(), key=lambda x: -x[1]), 11):
        ws[f"A{i}"] = src
        ws[f"B{i}"] = cnt
        ws[f"A{i}"].font = label_font
        ws[f"B{i}"].font = label_font
        ws[f"B{i}"].alignment = center
        fill = PatternFill("solid", start_color="EEF2FF") if i % 2 == 0 else PatternFill("solid", start_color="F8F9FF")
        ws[f"A{i}"].fill = fill
        ws[f"B{i}"].fill = fill
        ws.row_dimensions[i].height = 16

    # ── By State ──────────────────────────────────────────────────────────────
    from filter import get_state_label
    state_counts = Counter(get_state_label(j) or "Other" for j in jobs)
    ws["C9"] = "Jobs by State"
    ws["C9"].font = Font(name="Calibri", size=11, bold=True, color="1A237E")

    ws["C10"] = "State"
    ws["D10"] = "Count"
    for cell in [ws["C10"], ws["D10"]]:
        cell.font = heading_font
        cell.fill = hfill
        cell.alignment = center

    for i, (state, cnt) in enumerate(sorted(state_counts.items(), key=lambda x: -x[1]), 11):
        ws[f"C{i}"] = state
        ws[f"D{i}"] = cnt
        ws[f"C{i}"].font = label_font
        ws[f"D{i}"].font = label_font
        ws[f"D{i}"].alignment = center
        fill = PatternFill("solid", start_color="EEF2FF") if i % 2 == 0 else PatternFill("solid", start_color="F8F9FF")
        ws[f"C{i}"].fill = fill
        ws[f"D{i}"].fill = fill
        ws.row_dimensions[i].height = 16

    # ── Top companies ─────────────────────────────────────────────────────────
    co_counts = Counter(j.get("company_name", "") for j in jobs).most_common(15)
    ws["E9"] = "Top Companies"
    ws["E9"].font = Font(name="Calibri", size=11, bold=True, color="1A237E")

    ws["E10"] = "Company"
    ws["F10"] = "Listings"
    for cell in [ws["E10"], ws["F10"]]:
        cell.font = heading_font
        cell.fill = hfill
        cell.alignment = center

    for i, (co, cnt) in enumerate(co_counts, 11):
        ws[f"E{i}"] = co
        ws[f"F{i}"] = cnt
        ws[f"E{i}"].font = label_font
        ws[f"F{i}"].font = label_font
        ws[f"F{i}"].alignment = center
        fill = PatternFill("solid", start_color="EEF2FF") if i % 2 == 0 else PatternFill("solid", start_color="F8F9FF")
        ws[f"E{i}"].fill = fill
        ws[f"F{i}"].fill = fill
        ws.row_dimensions[i].height = 16

    return ws


def export_jobs_to_excel(jobs):
    manual_data = {}
    if os.path.exists(EXCEL_FILE):
        try:
            old_wb = load_workbook(EXCEL_FILE)
            sheet = old_wb["Jobs"] if "Jobs" in old_wb.sheetnames else old_wb.active
            manual_data = load_existing_manual_data(sheet)
        except BadZipFile:
            print(f"{EXCEL_FILE} is corrupt — recreating.")
        except Exception as e:
            print(f"Could not read existing workbook: {e}")

    wb = Workbook()
    _write_jobs_sheet(wb, jobs, manual_data)
    _write_stats_sheet(wb, jobs)

    try:
        wb.save(EXCEL_FILE)
        print(f"Saved {EXCEL_FILE} ({len(jobs)} jobs)")
    except PermissionError:
        print(f"Could not save {EXCEL_FILE} — close it in Excel first.")
